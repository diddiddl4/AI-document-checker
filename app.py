import streamlit as st
import os
import openpyxl
from openpyxl.utils import get_column_letter
from copy import copy
from docx import Document
from pptx import Presentation
import PyPDF2
from datetime import datetime
import io
import base64
from anthropic import Anthropic
from pdf2image import convert_from_bytes
from PIL import Image

# 페이지 설정
st.set_page_config(
    page_title="AI 문서 점검기 - 경원알미늄",
    page_icon="🔍",
    layout="wide"
)

# CSS 스타일
st.markdown("""
<style>
    .main {
        background: linear-gradient(135deg, #2193b0 0%, #6dd5ed 100%);
    }
    .stApp {
        background: linear-gradient(135deg, #2193b0 0%, #6dd5ed 100%);
    }
    h1 {
        color: white;
        text-align: center;
    }
    .footer {
        position: fixed;
        bottom: 0;
        left: 0;
        right: 0;
        background: linear-gradient(135deg, #2193b0 0%, #6dd5ed 100%);
        color: white;
        text-align: center;
        padding: 10px;
    }
</style>
""", unsafe_allow_html=True)

# Claude API 초기화
def get_claude_client():
    api_key = st.secrets.get('ANTHROPIC_API_KEY') or os.environ.get('ANTHROPIC_API_KEY')
    if not api_key:
        return None
    return Anthropic(api_key=api_key)

class DocumentAnalyzer:
    def __init__(self, filepath, mode='standard'):
        self.filepath = filepath
        self.file_ext = os.path.splitext(filepath)[1].lower()
        self.mode = mode
        self.issues = []
        self.warnings = []
        self.score = 100
        self.cell_issues = []
        self.ocr_text = None
        
    def analyze(self):
        if self.file_ext in ['.xlsx', '.xls']:
            return self._analyze_excel()
        elif self.file_ext in ['.docx', '.doc']:
            return self._analyze_word()
        elif self.file_ext in ['.pptx', '.ppt']:
            return self._analyze_ppt()
        elif self.file_ext == '.pdf':
            return self._analyze_pdf()
        elif self.file_ext in ['.jpg', '.jpeg', '.png', '.gif', '.bmp']:
            return self._analyze_image()
        return self._get_result()
    
    def ocr_with_claude(self, image_data, is_bytes=True):
        """Claude API로 이미지 OCR"""
        try:
            client = get_claude_client()
            if not client:
                return "API 키가 설정되지 않았습니다."
            
            # 이미지를 base64로 변환
            if is_bytes:
                image_b64 = base64.b64encode(image_data).decode('utf-8')
            else:
                with open(image_data, 'rb') as f:
                    image_b64 = base64.b64encode(f.read()).decode('utf-8')
            
            # Claude API 호출
            message = client.messages.create(
                model="claude-sonnet-4-20250514",  # 최신 모델명
                max_tokens=4096,
                messages=[{
                    "role": "user",
                    "content": [
                        {
                            "type": "image",
                            "source": {
                                "type": "base64",
                                "media_type": "image/png",
                                "data": image_b64
                            }
                        },
                        {
                            "type": "text",
                            "text": """이 이미지의 모든 텍스트를 정확하게 추출해주세요. 

요구사항:
1. 원본의 단락 구분을 정확히 유지
2. 적절한 띄어쓰기 적용
3. 표가 있다면 마크다운 표 형식으로
4. 제목과 본문 구분 명확히
5. 불필요한 공백 제거

깔끔하고 읽기 쉬운 형식으로 작성해주세요."""
                        }
                    ]
                }]
            )
            
            return message.content[0].text
            
        except Exception as e:
            st.error(f"🐛 OCR 오류: {str(e)}")
            import traceback
            st.code(traceback.format_exc())
            return f"OCR 오류: {str(e)}"
    
    def _analyze_excel(self):
        try:
            wb = openpyxl.load_workbook(self.filepath, data_only=False)
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                
                # 병합 셀 검사
                merged_cells = list(sheet.merged_cells.ranges)
                if merged_cells:
                    self.score -= len(merged_cells) * 3
                    for merged in merged_cells:
                        self.cell_issues.append({
                            'sheet': sheet_name,
                            'cell': str(merged),
                            'type': 'MERGED_CELL',
                            'severity': 'HIGH',
                            'message': f'병합된 셀: {merged}',
                            'recommendation': '병합 해제 후 데이터 정규화 필요'
                        })
                    self.issues.append({
                        'type': 'MERGED_CELLS',
                        'count': len(merged_cells),
                        'message': f'{len(merged_cells)}개의 병합 셀 발견'
                    })
                
                # 줄바꿈 검사
                newline_count = 0
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str) and '\n' in cell.value:
                            newline_count += 1
                
                if newline_count > 0:
                    self.warnings.append({
                        'type': 'NEWLINES',
                        'count': newline_count,
                        'message': f'{newline_count}개 셀에 줄바꿈 포함'
                    })
                
                # 숨겨진 행/열
                hidden_rows = [i for i in range(1, sheet.max_row + 1) if sheet.row_dimensions[i].hidden]
                hidden_cols = [i for i in range(1, sheet.max_column + 1) 
                              if sheet.column_dimensions[get_column_letter(i)].hidden]
                
                if hidden_rows or hidden_cols:
                    self.score -= 15
                    self.issues.append({
                        'type': 'HIDDEN_DATA',
                        'message': f'숨겨진 행 {len(hidden_rows)}개, 열 {len(hidden_cols)}개'
                    })
            
        except Exception as e:
            self.issues.append({'type': 'ERROR', 'message': str(e)})
        
        return self._get_result()
    
    def _analyze_word(self):
        try:
            doc = Document(self.filepath)
            table_count = len(doc.tables)
            if table_count > 0:
                self.warnings.append({
                    'type': 'TABLES',
                    'message': f'{table_count}개의 표 발견'
                })
        except Exception as e:
            self.issues.append({'type': 'ERROR', 'message': str(e)})
        return self._get_result()
    
    def _analyze_ppt(self):
        try:
            prs = Presentation(self.filepath)
            slide_count = len(prs.slides)
            if slide_count > 50:
                self.score -= 10
                self.warnings.append({
                    'type': 'MANY_SLIDES',
                    'message': f'{slide_count}개의 슬라이드'
                })
        except Exception as e:
            self.issues.append({'type': 'ERROR', 'message': str(e)})
        return self._get_result()
    
    def _analyze_pdf(self):
        """PDF 분석 + OCR"""
        try:
            pdf = PyPDF2.PdfReader(self.filepath)
            text_extractable = False
            extracted_text = ""
            
            # 일반 텍스트 추출 시도
            for page in pdf.pages[:3]:
                text = page.extract_text().strip()
                if text:
                    text_extractable = True
                    extracted_text += text + "\n\n"
            
            if not text_extractable:
                # 스캔 PDF → OCR 실행
                self.score -= 20
                self.issues.append({
                    'type': 'SCANNED_PDF',
                    'message': '스캔된 PDF - OCR 처리 중...'
                })
                
                client = get_claude_client()
                if client:
                    with st.spinner('📸 Claude AI로 텍스트 추출 중...'):
                        # PDF를 이미지로 변환
                        with open(self.filepath, 'rb') as f:
                            images = convert_from_bytes(f.read(), first_page=1, last_page=3)
                        
                        full_text = ""
                        for i, image in enumerate(images):
                            # PIL Image를 bytes로 변환
                            img_byte_arr = io.BytesIO()
                            image.save(img_byte_arr, format='PNG')
                            img_byte_arr = img_byte_arr.getvalue()
                            
                            # OCR 실행
                            page_text = self.ocr_with_claude(img_byte_arr, is_bytes=True)
                            full_text += f"\n\n=== 페이지 {i+1} ===\n\n{page_text}"
                        
                        self.ocr_text = full_text
                        self.warnings.append({
                            'type': 'OCR_SUCCESS',
                            'message': f'Claude OCR로 {len(images)}페이지 텍스트 추출 완료'
                        })
                else:
                    self.warnings.append({
                        'type': 'NO_API_KEY',
                        'message': 'OCR을 위해 API 키가 필요합니다'
                    })
            else:
                self.ocr_text = extracted_text
            
        except Exception as e:
            self.issues.append({'type': 'ERROR', 'message': str(e)})
            st.error(f"PDF 분석 오류: {str(e)}")
        return self._get_result()
    
    def _analyze_image(self):
        """이미지 분석 + OCR"""
        try:
            client = get_claude_client()
            if not client:
                self.score = 50
                self.warnings.append({
                    'type': 'NO_API_KEY',
                    'message': 'OCR을 위해 API 키가 필요합니다'
                })
                return self._get_result()
            
            with st.spinner('📸 Claude AI로 텍스트 추출 중...'):
                # OCR 실행
                extracted_text = self.ocr_with_claude(self.filepath, is_bytes=False)
                
                if extracted_text and not extracted_text.startswith("OCR 오류"):
                    self.score = 75
                    self.warnings.append({
                        'type': 'IMAGE_OCR',
                        'message': 'Claude OCR로 텍스트 추출 완료'
                    })
                    self.ocr_text = extracted_text
                else:
                    self.score = 30
                    self.issues.append({
                        'type': 'OCR_FAILED',
                        'message': '텍스트 추출 실패'
                    })
        except Exception as e:
            self.issues.append({'type': 'ERROR', 'message': str(e)})
            st.error(f"이미지 분석 오류: {str(e)}")
        return self._get_result()
    
    def _get_result(self):
        self.score = max(0, min(100, self.score))
        
        if self.score >= 80:
            grade = 'A'
        elif self.score >= 60:
            grade = 'B'
        elif self.score >= 40:
            grade = 'C'
        else:
            grade = 'D'
        
        return {
            'score': self.score,
            'grade': grade,
            'issues': self.issues,
            'warnings': self.warnings,
            'cell_issues': self.cell_issues,
            'file_type': self.file_ext,
            'mode': self.mode
        }
    
    def generate_optimized_version(self):
        if self.file_ext not in ['.xlsx', '.xls']:
            return None
        
        try:
            wb = openpyxl.load_workbook(self.filepath)
            output = io.BytesIO()
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                
                # 병합 셀 해제 + 값 복사 + 서식 유지
                merged_ranges = list(sheet.merged_cells.ranges)
                for merged in merged_ranges:
                    min_col, min_row, max_col, max_row = merged.bounds
                    
                    source_cell = sheet.cell(min_row, min_col)
                    merged_value = source_cell.value
                    
                    # copy() 함수 사용
                    source_font = copy(source_cell.font) if source_cell.font else None
                    source_fill = copy(source_cell.fill) if source_cell.fill else None
                    source_border = copy(source_cell.border) if source_cell.border else None
                    source_alignment = copy(source_cell.alignment) if source_cell.alignment else None
                    
                    sheet.unmerge_cells(str(merged))
                    
                    for row in range(min_row, max_row + 1):
                        for col in range(min_col, max_col + 1):
                            cell = sheet.cell(row, col)
                            cell.value = merged_value
                            if source_font:
                                cell.font = copy(source_font)
                            if source_fill:
                                cell.fill = copy(source_fill)
                            if source_border:
                                cell.border = copy(source_border)
                            if source_alignment:
                                cell.alignment = copy(source_alignment)
                
                # 줄바꿈 제거
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            cell.value = cell.value.replace('\n', ' ')
                
                # 기호 변환 (분석 모드)
                if self.mode == 'analysis':
                    for row_idx, row in enumerate(sheet.iter_rows(), 1):
                        for col_idx, cell in enumerate(row, 1):
                            if row_idx > 1:
                                header_cell = sheet.cell(4, col_idx)
                                header = str(header_cell.value or '')
                                
                                if '여부' in header or '수령' in header:
                                    if cell.value in ['○', 'O', 'o', '●']:
                                        cell.value = '예'
                                    elif cell.value in ['', None, 'X', '×']:
                                        cell.value = '아니오'
                
                # 숨김 해제
                for i in range(1, sheet.max_row + 1):
                    sheet.row_dimensions[i].hidden = False
                for i in range(1, sheet.max_column + 1):
                    sheet.column_dimensions[get_column_letter(i)].hidden = False
            
            wb.save(output)
            output.seek(0)
            return output
            
        except Exception as e:
            st.error(f"최적화 오류: {e}")
            return None

# 메인 앱
st.title("🔍 AI 문서 점검기 Pro")
st.markdown("### 경원알미늄 - 탁월한 업무 시스템 구축 TFT")

# 모드 선택
col1, col2 = st.columns(2)
with col1:
    mode = st.radio(
        "최적화 모드",
        ["표준 모드", "분석 모드"],
        help="표준: 병합셀 해제 + 줄바꿈 제거 | 분석: 표준 + 기호변환"
    )

selected_mode = 'standard' if mode == "표준 모드" else 'analysis'

# 파일 업로드
uploaded_file = st.file_uploader(
    "파일을 선택하세요",
    type=['xlsx', 'xls', 'docx', 'doc', 'pptx', 'ppt', 'pdf', 'jpg', 'jpeg', 'png'],
    help="Excel, Word, PowerPoint, PDF, 이미지 지원"
)

if uploaded_file:
    # 임시 파일 저장
    with open(f"temp_{uploaded_file.name}", "wb") as f:
        f.write(uploaded_file.getbuffer())
    
    # 분석
    with st.spinner('분석 중...'):
        analyzer = DocumentAnalyzer(f"temp_{uploaded_file.name}", mode=selected_mode)
        result = analyzer.analyze()
    
    # 결과 표시
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.metric("점수", f"{result['score']}점")
    with col2:
        st.metric("등급", result['grade'])
    with col3:
        st.metric("모드", "표준" if selected_mode == 'standard' else "분석")
    
    # 이슈 표시
    if result['issues']:
        st.subheader("🚨 주요 이슈")
        for issue in result['issues']:
            st.error(f"**{issue.get('type')}**: {issue.get('message')}")
    
    if result['warnings']:
        st.subheader("⚠️ 경고")
        for warning in result['warnings']:
            st.warning(f"**{warning.get('type')}**: {warning.get('message')}")
    
    if result['cell_issues']:
        st.subheader("📍 셀별 문제점")
        for cell_issue in result['cell_issues'][:10]:
            st.info(f"{cell_issue['sheet']} - {cell_issue['cell']}: {cell_issue['message']}")
    
    # 다운로드 버튼
    st.subheader("📥 다운로드")
    
    col1, col2 = st.columns(2)
    
    with col1:
        if result['file_type'] in ['.xlsx', '.xls']:
            optimized = analyzer.generate_optimized_version()
            if optimized:
                st.download_button(
                    label="✨ AI 최적화 버전",
                    data=optimized,
                    file_name=f"AI최적화_{uploaded_file.name}",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
    
    with col2:
        if analyzer.ocr_text:
            st.download_button(
                label="📝 OCR 텍스트 추출",
                data=analyzer.ocr_text,
                file_name=f"OCR_{uploaded_file.name}.txt",
                mime="text/plain"
            )
    
    # OCR 결과 미리보기
    if analyzer.ocr_text:
        with st.expander("👁️ 추출된 텍스트 미리보기"):
            st.text_area("", analyzer.ocr_text, height=300)
    
    # 임시 파일 삭제
    try:
        os.remove(f"temp_{uploaded_file.name}")
    except:
        pass

# 푸터
st.markdown("""
<div class="footer">
경원알미늄 - 탁월한 업무 시스템 구축 TFT
</div>
""", unsafe_allow_html=True)
